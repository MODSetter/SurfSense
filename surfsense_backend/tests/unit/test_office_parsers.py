"""Unit tests for native Office file parsers (no DB, no external services)."""

import tempfile

import pytest
from openpyxl import Workbook

pytestmark = pytest.mark.unit


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _create_xlsx(sheets: dict[str, list[list]]) -> str:
    """Create a real .xlsx file on disk and return its path.

    ``sheets`` maps sheet name -> list of rows, where each row is a list of
    cell values.
    """
    wb = Workbook()
    first = True
    for name, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet(title=name)
        if first:
            ws.title = name
            first = False
        for row in rows:
            ws.append(row)
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    wb.close()
    tmp.close()
    return tmp.name


# ---------------------------------------------------------------------------
# Tracer bullet: cell values appear in markdown
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_parse_excel_produces_markdown_with_cell_values():
    """A single-sheet .xlsx with known data produces markdown containing those values."""
    from app.utils.office_parsers import parse_excel_to_markdown

    path = _create_xlsx(
        {"Sales": [["Product", "Revenue"], ["Widget", 1500], ["Gadget", 3200]]}
    )

    md = await parse_excel_to_markdown(path, filename="report.xlsx")

    assert "Product" in md
    assert "Revenue" in md
    assert "Widget" in md
    assert "1500" in md
    assert "Gadget" in md
    assert "3200" in md
    assert "report.xlsx" in md
    assert "|" in md


# ---------------------------------------------------------------------------
# Multi-sheet workbooks include all sheets
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_parse_excel_includes_all_sheets():
    """Both sheet names and their data appear in the output."""
    from app.utils.office_parsers import parse_excel_to_markdown

    path = _create_xlsx(
        {
            "Inventory": [["Item", "Qty"], ["Bolts", 200]],
            "Pricing": [["Item", "Price"], ["Bolts", 4.50]],
        }
    )

    md = await parse_excel_to_markdown(path, filename="multi.xlsx")

    assert "Inventory" in md
    assert "Pricing" in md
    assert "Bolts" in md
    assert "200" in md
    assert "4.5" in md


# ---------------------------------------------------------------------------
# Empty spreadsheet raises ValueError
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_parse_excel_raises_on_empty_file():
    """An .xlsx with no data raises ValueError."""
    from app.utils.office_parsers import parse_excel_to_markdown

    wb = Workbook()
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    wb.close()
    tmp.close()

    with pytest.raises(ValueError, match="No data found"):
        await parse_excel_to_markdown(tmp.name, filename="empty.xlsx")


# ---------------------------------------------------------------------------
# _parse_file_to_markdown routes .xlsx natively (no ETL call)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_parse_file_to_markdown_routes_xlsx_natively():
    """content_extractor._parse_file_to_markdown uses native parser for .xlsx."""
    from app.connectors.google_drive.content_extractor import _parse_file_to_markdown

    path = _create_xlsx(
        {"Data": [["Name", "Score"], ["Alice", 95], ["Bob", 82]]}
    )

    md = await _parse_file_to_markdown(path, "grades.xlsx")

    assert "Alice" in md
    assert "95" in md
    assert "Bob" in md
    assert "82" in md
