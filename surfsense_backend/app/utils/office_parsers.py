"""Native parsers for Office file formats."""

import asyncio
import logging
import threading
import time
from pathlib import Path

logger = logging.getLogger(__name__)

EXCEL_EXTENSIONS = (".xlsx",)


def _parse_excel_sync(file_path: str) -> str:
    """Parse an .xlsx file into markdown tables (synchronous)."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    markdown_parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        non_empty_rows = [r for r in rows if any(c is not None for c in r)]
        if not non_empty_rows:
            continue

        markdown_parts.append(f"## {sheet_name}\n")
        max_cols = max(len(row) for row in non_empty_rows)

        header = non_empty_rows[0]
        hdr = [str(c if c is not None else "") for c in header]
        hdr.extend([""] * (max_cols - len(hdr)))
        markdown_parts.append("| " + " | ".join(hdr) + " |")
        markdown_parts.append("| " + " | ".join("---" for _ in range(max_cols)) + " |")

        for row in non_empty_rows[1:]:
            cells = [str(c if c is not None else "") for c in row]
            cells.extend([""] * (max_cols - len(cells)))
            markdown_parts.append("| " + " | ".join(cells) + " |")

        markdown_parts.append("")

    wb.close()
    return "\n".join(markdown_parts)


async def parse_excel_to_markdown(file_path: str, filename: str = "") -> str:
    """Parse an .xlsx file into markdown tables (async wrapper).

    Raises ``ValueError`` if no data is found in the workbook.
    """
    t0 = time.monotonic()
    logger.info(
        "[excel-parse] START file=%s thread=%s",
        filename,
        threading.current_thread().name,
    )

    result = await asyncio.to_thread(_parse_excel_sync, file_path)

    logger.info(
        "[excel-parse] END file=%s elapsed=%.2fs",
        filename,
        time.monotonic() - t0,
    )

    if not result.strip():
        raise ValueError(f"No data found in Excel file: {filename or file_path}")

    title = f"# {filename}\n\n" if filename else ""
    return title + result
